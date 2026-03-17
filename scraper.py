#!/usr/bin/env python3
"""
Google Maps Business Data Scraper
Extracts Name, Address, Category, Website, Email, and Phone
from Google Maps using browser automation — no API key or login required.

Author    : Soubarna Karmakar
Copyright : © 2025 Soubarna Karmakar. All rights reserved.
Version   : 2.0
"""

__author__    = "Soubarna Karmakar"
__copyright__ = "Copyright © 2025 Soubarna Karmakar. All rights reserved."
__version__   = "2.0"

# ── Standard Library ──────────────────────────────────────────────────────────
import os, re, sys, time, queue, random, threading
from datetime import datetime
from urllib.parse import quote_plus, urljoin

# ── Third-Party: UI ───────────────────────────────────────────────────────────
import customtkinter as ctk
import tkinter as tk
from tkinter import ttk, filedialog, messagebox

# ── Third-Party: Data ─────────────────────────────────────────────────────────
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Third-Party: Browser ──────────────────────────────────────────────────────
from playwright.sync_api import sync_playwright, TimeoutError as PWTimeout

# ── Third-Party: Email Extraction (optional) ──────────────────────────────────
try:
    import requests
    from bs4 import BeautifulSoup
    _EMAIL_LIBS = True
except ImportError:
    _EMAIL_LIBS = False

# ══════════════════════════════════════════════════════════════════════════════
# CONSTANTS
# ══════════════════════════════════════════════════════════════════════════════
APP_NAME    = "Google Maps Scraper"
APP_VERSION = "2.0"

# Palette
C_ACCENT   = "#1E88E5"
C_ACCENT2  = "#1565C0"
C_SUCCESS  = "#43A047"
C_ERROR    = "#E53935"
C_WARN     = "#FB8C00"
C_BG       = "#12122A"
C_PANEL    = "#1A1A38"
C_CARD     = "#1E2040"
C_HEADER   = "#0D47A1"
C_FG       = "#E8E8FF"
C_MUTED    = "#7B7B9E"

DATA_COLS  = ["Name", "Address", "Category", "Phone", "Website", "Email", "Query"]

# Human-like timing ranges (seconds)
T_PAGE    = (2.5, 4.0)
T_SCROLL  = (1.5, 2.5)
T_PLACE   = (1.8, 3.0)
T_SITE    = 8          # timeout for website email fetch

# ══════════════════════════════════════════════════════════════════════════════
# STEALTH SCRIPT  (injected into every new page)
# ══════════════════════════════════════════════════════════════════════════════
STEALTH_JS = """
() => {
    Object.defineProperty(navigator, 'webdriver',  { get: () => undefined });
    Object.defineProperty(navigator, 'plugins',    { get: () => [1, 2, 3, 4, 5] });
    Object.defineProperty(navigator, 'languages',  { get: () => ['en-US', 'en'] });
    Object.defineProperty(navigator, 'platform',   { get: () => 'Win32' });
    window.chrome = { runtime: {} };
}
"""

# ══════════════════════════════════════════════════════════════════════════════
# EMAIL EXTRACTOR
# ══════════════════════════════════════════════════════════════════════════════
class EmailExtractor:
    _PATTERN = re.compile(r'[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}')
    _SKIP    = ('example', 'test', 'noreply', 'no-reply', 'sentry', 'wixpress',
                'yourdomain', 'domain.com', '@2x', '.png', '.jpg', '.gif',
                'placeholder', 'email.com', 'user@', 'support@example')

    def __init__(self):
        if not _EMAIL_LIBS:
            return
        self._session = requests.Session()
        self._session.headers.update({
            'User-Agent': ('Mozilla/5.0 (Windows NT 10.0; Win64; x64) '
                           'AppleWebKit/537.36 (KHTML, like Gecko) '
                           'Chrome/122.0.0.0 Safari/537.36'),
            'Accept-Language': 'en-US,en;q=0.9',
        })

    def extract(self, url: str) -> str:
        if not _EMAIL_LIBS or not url:
            return ""
        try:
            r = self._session.get(url, timeout=T_SITE, allow_redirects=True)
            r.raise_for_status()
            soup = BeautifulSoup(r.text, 'html.parser')

            # 1. mailto: links
            for a in soup.find_all('a', href=True):
                h = a['href']
                if h.lower().startswith('mailto:'):
                    email = h[7:].split('?')[0].strip()
                    if self._valid(email):
                        return email

            # 2. Page text scan
            emails = [e for e in self._PATTERN.findall(r.text) if self._valid(e)]
            if emails:
                return emails[0]

            # 3. Try /contact page
            contact = self._find_contact(soup, url)
            if contact:
                r2 = self._session.get(contact, timeout=T_SITE)
                emails2 = [e for e in self._PATTERN.findall(r2.text) if self._valid(e)]
                if emails2:
                    return emails2[0]
        except Exception:
            pass
        return ""

    def _valid(self, email: str) -> bool:
        el = email.lower()
        return not any(s in el for s in self._SKIP)

    def _find_contact(self, soup, base: str) -> str:
        for a in soup.find_all('a', href=True):
            text = a.get_text(strip=True).lower()
            href = a['href']
            if any(w in text for w in ('contact', 'about us', 'get in touch')):
                full = urljoin(base, href)
                if full != base:
                    return full
        return ""


# ══════════════════════════════════════════════════════════════════════════════
# GOOGLE MAPS SCRAPER ENGINE
# ══════════════════════════════════════════════════════════════════════════════
class MapsScraper:
    """Runs entirely inside a background thread; communicates via queues."""

    def __init__(self, headless: bool, extract_emails: bool):
        self.headless       = headless
        self.extract_emails = extract_emails
        self._stop          = threading.Event()
        self.result_q: queue.Queue = queue.Queue()
        self.status_q: queue.Queue = queue.Queue()
        self._email_ext     = EmailExtractor() if extract_emails else None

    # ── Public control ────────────────────────────────────────────────────────
    def stop(self):
        self._stop.set()

    # ── Internal helpers ──────────────────────────────────────────────────────
    def _push_status(self, msg: str, level: str = "info"):
        self.status_q.put({"msg": msg, "level": level})

    def _delay(self, rng=T_PAGE):
        time.sleep(random.uniform(*rng))

    def _dismiss_overlay(self, page):
        """Dismiss cookie / consent banners if present."""
        for sel in [
            "button:has-text('Accept all')",
            "button:has-text('Accept')",
            "button:has-text('Agree')",
            "button:has-text('I agree')",
            "[aria-label='Accept all']",
        ]:
            try:
                btn = page.query_selector(sel)
                if btn and btn.is_visible():
                    btn.click()
                    time.sleep(0.8)
                    return
            except Exception:
                pass

    # ── Phase 1: collect all place URLs from the search feed ──────────────────
    def _collect_urls(self, page, query: str) -> list:
        search_url = "https://www.google.com/maps/search/" + quote_plus(query)
        self._push_status(f'Searching: "{query}"')
        page.goto(search_url, wait_until="domcontentloaded")
        self._delay(T_PAGE)
        self._dismiss_overlay(page)

        # Wait for the results feed
        try:
            page.wait_for_selector('div[role="feed"]', timeout=15_000)
        except PWTimeout:
            self._push_status("Results feed not found — skipping query.", "warning")
            return []

        collected: list   = []
        seen:      set    = set()
        no_new_streak     = 0
        MAX_NO_NEW        = 4

        while not self._stop.is_set():
            # --- Grab all visible place links ---
            try:
                handles = page.query_selector_all('div[role="feed"] a[href*="/maps/place/"]')
            except Exception:
                break

            new_found = 0
            for h in handles:
                try:
                    href = h.get_attribute("href") or ""
                    # Normalise: keep everything up to the data= segment
                    base = href.split("?")[0]
                    if base not in seen:
                        seen.add(base)
                        collected.append(href)
                        new_found += 1
                except Exception:
                    pass

            if new_found:
                no_new_streak = 0
                self._push_status(f'Found {len(collected)} places so far…')
            else:
                no_new_streak += 1

            # --- Check for end-of-list text ---
            try:
                src = page.content()
                if any(p in src for p in [
                    "You've reached the end of the list",
                    "No results for",
                    "didn't match any places",
                    "We couldn't find",
                ]):
                    self._push_status(
                        f'End of results — {len(collected)} places found.', "success")
                    break
            except Exception:
                pass

            if no_new_streak >= MAX_NO_NEW:
                self._push_status(
                    f'No more new results — {len(collected)} places found.', "success")
                break

            # --- Scroll the feed ---
            try:
                page.evaluate("""
                    const feed = document.querySelector('div[role="feed"]');
                    if (feed) feed.scrollTop = feed.scrollHeight;
                """)
            except Exception:
                pass
            self._delay(T_SCROLL)

        return collected

    # ── Phase 2: extract data from a single place page ────────────────────────
    def _extract_place(self, page, url: str, query: str) -> dict | None:
        try:
            page.goto(url, wait_until="domcontentloaded")
            self._delay(T_PLACE)

            # Wait for the title
            try:
                page.wait_for_selector("h1", timeout=12_000)
            except PWTimeout:
                return None

            data = {c: "" for c in DATA_COLS}
            data["Query"] = query

            # ── Name ──────────────────────────────────────────────────────────
            for sel in ["h1.DUwDvf", "h1"]:
                el = page.query_selector(sel)
                if el:
                    data["Name"] = (el.text_content() or "").strip()
                    if data["Name"]:
                        break
            if not data["Name"]:
                return None

            # ── Category ──────────────────────────────────────────────────────
            for sel in [
                "button.DkEaL",
                "div.fontBodyMedium span.mgr77e",
                "[jsaction*='category']",
            ]:
                el = page.query_selector(sel)
                if el:
                    txt = (el.text_content() or "").strip()
                    if txt:
                        data["Category"] = txt
                        break

            # ── Info items (Address / Phone / Website) ────────────────────────
            items = page.query_selector_all("[data-item-id]")
            for item in items:
                try:
                    item_id = item.get_attribute("data-item-id") or ""
                    text    = (item.text_content() or "").strip()

                    if item_id == "address":
                        data["Address"] = text.replace("\n", ", ")

                    elif item_id.startswith("phone:tel:"):
                        # Text may contain icon label — keep first line
                        data["Phone"] = text.split("\n")[0]

                    elif item_id == "authority":
                        link = item.query_selector("a")
                        if link:
                            data["Website"] = link.get_attribute("href") or text
                        else:
                            data["Website"] = text
                except Exception:
                    pass

            # ── Fallback selectors ────────────────────────────────────────────
            if not data["Address"]:
                el = page.query_selector("button[data-item-id='address']")
                if el:
                    data["Address"] = (el.text_content() or "").strip().replace("\n", ", ")

            if not data["Phone"]:
                for sel in [
                    "[data-item-id^='phone:tel:']",
                    "button[aria-label*='Phone']",
                    "[data-tooltip*='phone']",
                ]:
                    el = page.query_selector(sel)
                    if el:
                        t = (el.text_content() or "").strip()
                        if not t:
                            t = (el.get_attribute("aria-label") or "").replace("Phone:", "").strip()
                        m = re.search(r'[\+\d][\d\s\-\(\)\.]{6,}', t)
                        if m:
                            data["Phone"] = m.group().strip()
                            break

            if not data["Website"]:
                el = page.query_selector("a[data-item-id='authority']")
                if el:
                    data["Website"] = el.get_attribute("href") or ""

            # ── Email (from website) ──────────────────────────────────────────
            if self.extract_emails and self._email_ext and data["Website"]:
                data["Email"] = self._email_ext.extract(data["Website"]) or "N/A"
            else:
                data["Email"] = ""

            return data

        except Exception as e:
            self._push_status(f"Extract error: {str(e)[:60]}", "warning")
            return None

    # ── Browser installation helper (works in frozen PyInstaller app too) ────
    def _ensure_browser(self) -> bool:
        """Return True if Chromium is ready; auto-install if missing."""
        def _install_cmd():
            """Return the correct command to run `playwright install chromium`."""
            if getattr(sys, "frozen", False):
                # Inside a PyInstaller bundle — use the bundled playwright driver
                base = sys._MEIPASS
                if sys.platform == "win32":
                    drv = os.path.join(base, "playwright", "driver", "playwright.cmd")
                elif sys.platform == "darwin":
                    drv = os.path.join(base, "playwright", "driver", "playwright.sh")
                else:
                    drv = os.path.join(base, "playwright", "driver", "playwright.sh")
                if os.path.exists(drv):
                    return [drv, "install", "chromium"]
            # Normal (non-frozen) run — use the Python interpreter
            return [sys.executable, "-m", "playwright", "install", "chromium"]

        import subprocess
        try:
            # Quick probe: attempt to get the executable path
            with sync_playwright() as pw:
                ep = pw.chromium.executable_path
            if ep and os.path.exists(ep):
                return True   # Already installed
        except Exception:
            pass

        # Not installed — run `playwright install chromium`
        self._push_status(
            "Downloading Chromium (~120 MB, one-time setup)…", "warning")
        try:
            subprocess.run(_install_cmd(), check=True, timeout=600)
            self._push_status("Chromium ready.", "success")
            return True
        except Exception as e:
            self._push_status(f"Browser install failed: {e}", "error")
            return False

    # ── Main entry (called from thread) ──────────────────────────────────────
    def run(self, queries: list, on_result, on_complete):
        self._stop.clear()
        total = 0

        try:
            self._push_status("Checking browser…")
            if not self._ensure_browser():
                self._push_status(
                    "Cannot start: Chromium not available. "
                    "Run: python -m playwright install chromium", "error")
                return

            self._push_status("Launching browser…")
            with sync_playwright() as pw:
                browser = pw.chromium.launch(
                    headless=self.headless,
                    args=[
                        "--disable-blink-features=AutomationControlled",
                        "--no-sandbox",
                        "--disable-notifications",
                        "--lang=en-US",
                    ],
                )
                ctx = browser.new_context(
                    user_agent=(
                        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                        "AppleWebKit/537.36 (KHTML, like Gecko) "
                        "Chrome/122.0.0.0 Safari/537.36"
                    ),
                    viewport={"width": 1366, "height": 768},
                    locale="en-US",
                    timezone_id="America/New_York",
                )
                ctx.add_init_script(STEALTH_JS)
                page = ctx.new_page()
                self._push_status("Browser ready.", "success")

                for qi, query in enumerate(queries, 1):
                    query = query.strip()
                    if not query or self._stop.is_set():
                        continue

                    self._push_status(f"[{qi}/{len(queries)}] Collecting results for: \"{query}\"")
                    urls = self._collect_urls(page, query)

                    if not urls:
                        self._push_status(f'No places found for "{query}"', "warning")
                        continue

                    self._push_status(f"Extracting data from {len(urls)} places…")
                    for pi, url in enumerate(urls, 1):
                        if self._stop.is_set():
                            break
                        self._push_status(
                            f'[{qi}/{len(queries)}] "{query[:35]}…" → place {pi}/{len(urls)}'
                        )
                        d = self._extract_place(page, url, query)
                        if d:
                            total += 1
                            on_result(d, total)
                        self._delay((0.4, 1.2))

                browser.close()

            if not self._stop.is_set():
                self._push_status(
                    f"Done! {total} record{'s' if total != 1 else ''} extracted.", "success")
            else:
                self._push_status(f"Stopped. {total} records saved so far.", "warning")

        except Exception as e:
            self._push_status(f"Fatal error: {e}", "error")
        finally:
            on_complete()


# ══════════════════════════════════════════════════════════════════════════════
# EXCEL EXPORTER
# ══════════════════════════════════════════════════════════════════════════════
class ExcelExporter:
    _HDR_FILL  = PatternFill("solid", fgColor="0D47A1")
    _HDR_FONT  = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    _HDR_ALIGN = Alignment(horizontal="center", vertical="center")
    _ALT_FILL  = PatternFill("solid", fgColor="E3F2FD")
    _BORDER    = Border(
        left   = Side(style="thin", color="BDBDBD"),
        right  = Side(style="thin", color="BDBDBD"),
        top    = Side(style="thin", color="BDBDBD"),
        bottom = Side(style="thin", color="BDBDBD"),
    )
    _COL_WIDTHS = {
        "Name": 32, "Address": 42, "Category": 22,
        "Phone": 18, "Website": 38, "Email": 32, "Query": 28,
    }

    def export(self, records: list, path: str) -> str:
        wb = openpyxl.Workbook()

        # ── Data sheet ────────────────────────────────────────────────────────
        ws = wb.active
        ws.title = "Maps Data"

        for ci, col in enumerate(DATA_COLS, 1):
            c = ws.cell(1, ci, col)
            c.font, c.fill, c.alignment, c.border = (
                self._HDR_FONT, self._HDR_FILL, self._HDR_ALIGN, self._BORDER)

        for ri, row in enumerate(records, 2):
            for ci, col in enumerate(DATA_COLS, 1):
                c = ws.cell(ri, ci, row.get(col, ""))
                c.border    = self._BORDER
                c.alignment = Alignment(vertical="center")
                if ri % 2 == 0:
                    c.fill = self._ALT_FILL

        for ci, col in enumerate(DATA_COLS, 1):
            ws.column_dimensions[get_column_letter(ci)].width = self._COL_WIDTHS.get(col, 20)

        ws.row_dimensions[1].height = 26
        for ri in range(2, len(records) + 2):
            ws.row_dimensions[ri].height = 18

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(DATA_COLS))}{len(records)+1}"

        # ── Summary sheet ─────────────────────────────────────────────────────
        ws2 = wb.create_sheet("Summary")
        ws2["A1"] = APP_NAME
        ws2["A1"].font = Font(bold=True, size=14, color="0D47A1")
        ws2["A2"] = f"Version {APP_VERSION}"
        ws2["A2"].font = Font(italic=True, color="888888")
        ws2["A4"] = "Exported On"
        ws2["B4"] = datetime.now().strftime("%Y-%m-%d  %H:%M:%S")
        ws2["A5"] = "Total Records"
        ws2["B5"] = len(records)
        ws2["A5"].font = ws2["A4"].font = Font(bold=True)

        ws2["A7"] = "Query Breakdown"
        ws2["A7"].font = Font(bold=True, size=11)
        ws2["A8"] = "Query"
        ws2["B8"] = "Count"
        ws2["A8"].font = ws2["B8"].font = Font(bold=True)

        qc: dict = {}
        for r in records:
            q = r.get("Query", "—")
            qc[q] = qc.get(q, 0) + 1
        for i, (q, cnt) in enumerate(qc.items(), 9):
            ws2[f"A{i}"] = q
            ws2[f"B{i}"] = cnt

        ws2.column_dimensions["A"].width = 45
        ws2.column_dimensions["B"].width = 10

        wb.save(path)
        return path


# ══════════════════════════════════════════════════════════════════════════════
# UI — MAIN APPLICATION
# ══════════════════════════════════════════════════════════════════════════════
class App(ctk.CTk):

    # ── Init ──────────────────────────────────────────────────────────────────
    def __init__(self):
        super().__init__()
        ctk.set_appearance_mode("dark")
        ctk.set_default_color_theme("blue")

        self.title(f"{APP_NAME}  v{APP_VERSION}")
        self.geometry("1380x820")
        self.minsize(1100, 660)

        self._scraper:  MapsScraper | None = None
        self._thread:   threading.Thread  | None = None
        self._running   = False
        self._records:  list = []
        self._sort_asc: dict = {c: True for c in DATA_COLS}

        self._build_ui()
        self._poll()

    # ═════════════════════════════════════════════════════════════════════════
    # UI CONSTRUCTION
    # ═════════════════════════════════════════════════════════════════════════

    def _build_ui(self):
        self.grid_columnconfigure(1, weight=1)
        self.grid_rowconfigure(0, weight=1)
        self._build_sidebar()
        self._build_content()

    # ── Sidebar ───────────────────────────────────────────────────────────────
    def _build_sidebar(self):
        sb = ctk.CTkFrame(self, width=295, corner_radius=0, fg_color=C_PANEL)
        sb.grid(row=0, column=0, sticky="nsew")
        sb.grid_propagate(False)
        sb.grid_columnconfigure(0, weight=1)

        # Title banner
        banner = ctk.CTkFrame(sb, fg_color=C_HEADER, corner_radius=0)
        banner.grid(row=0, column=0, sticky="ew")
        banner.grid_columnconfigure(0, weight=1)
        ctk.CTkLabel(banner, text="🗺  Maps Scraper",
                     font=ctk.CTkFont(size=19, weight="bold"),
                     text_color="white").grid(row=0, padx=16, pady=(14, 4))
        ctk.CTkLabel(banner, text="No API  •  No Login  •  No Limits",
                     font=ctk.CTkFont(size=10), text_color="#90CAF9"
                     ).grid(row=1, padx=16, pady=(0, 12))

        # Queries
        self._sb_label(sb, 1, "Search Queries")
        self._sb_hint(sb, 2, "One query per line — runs all simultaneously")
        self.txt_queries = ctk.CTkTextbox(
            sb, height=160, font=ctk.CTkFont(size=12), wrap="word",
            border_width=1, border_color="#2A2A5A")
        self.txt_queries.grid(row=3, column=0, padx=14, pady=(0, 4), sticky="ew")
        self.txt_queries.insert("1.0",
            "IT Companies in Koramangala\nIT Companies in Whitefield")

        self._divider(sb, 4)

        # Settings
        self._sb_label(sb, 5, "Settings")
        self.var_headless = ctk.BooleanVar(value=False)
        self.var_emails   = ctk.BooleanVar(value=False)

        ctk.CTkCheckBox(sb, text="Run browser in background (headless)",
                        variable=self.var_headless,
                        font=ctk.CTkFont(size=12)
                        ).grid(row=6, column=0, padx=16, pady=3, sticky="w")

        email_cb = ctk.CTkCheckBox(sb, text="Extract emails from websites",
                        variable=self.var_emails,
                        font=ctk.CTkFont(size=12))
        email_cb.grid(row=7, column=0, padx=16, pady=3, sticky="w")
        if not _EMAIL_LIBS:
            email_cb.configure(state="disabled",
                               text="Extract emails (install requests+bs4)")

        self._divider(sb, 8)

        # Action buttons
        self.btn_start = self._btn(sb, 9,  "▶   Start Scraping",
                                   C_SUCCESS, "#2E7D32", self._start)
        self.btn_stop  = self._btn(sb, 10, "⬛   Stop",
                                   C_ERROR,   "#B71C1C", self._stop,
                                   state="disabled")
        self.btn_export= self._btn(sb, 11, "📥   Export to Excel",
                                   C_ACCENT,  C_ACCENT2, self._export,
                                   state="disabled")
        self.btn_clear = self._btn(sb, 12, "🗑   Clear Results",
                                   "#37474F", "#263238", self._clear)

        self._divider(sb, 13)

        # Stats cards
        stats = ctk.CTkFrame(sb, fg_color=C_CARD, corner_radius=8)
        stats.grid(row=14, column=0, padx=14, pady=6, sticky="ew")
        stats.grid_columnconfigure((0, 1), weight=1)

        for col, label, color, attr in [
            (0, "Total Found",    C_ACCENT,  "lbl_total"),
            (1, "With Website",   C_SUCCESS, "lbl_web"),
        ]:
            ctk.CTkLabel(stats, text=label,
                         font=ctk.CTkFont(size=10), text_color=C_MUTED
                         ).grid(row=0, column=col, padx=8, pady=(8, 0))
            lbl = ctk.CTkLabel(stats, text="0",
                               font=ctk.CTkFont(size=26, weight="bold"),
                               text_color=color)
            lbl.grid(row=1, column=col, padx=8, pady=(0, 10))
            setattr(self, attr, lbl)

        # Footer
        ctk.CTkLabel(sb, text=f"v{APP_VERSION}  •  Playwright  •  CustomTkinter",
                     font=ctk.CTkFont(size=9), text_color=C_MUTED
                     ).grid(row=20, column=0, padx=14, pady=10, sticky="s")
        sb.grid_rowconfigure(19, weight=1)

    def _sb_label(self, parent, row, text):
        ctk.CTkLabel(parent, text=text,
                     font=ctk.CTkFont(size=13, weight="bold"),
                     anchor="w").grid(row=row, column=0, padx=16, pady=(12, 3), sticky="ew")

    def _sb_hint(self, parent, row, text):
        ctk.CTkLabel(parent, text=text,
                     font=ctk.CTkFont(size=10), text_color=C_MUTED,
                     anchor="w").grid(row=row, column=0, padx=16, pady=(0, 4), sticky="ew")

    def _divider(self, parent, row):
        ctk.CTkFrame(parent, height=1, fg_color="#2A2A55"
                     ).grid(row=row, column=0, padx=14, pady=6, sticky="ew")

    def _btn(self, parent, row, text, fg, hover, cmd, state="normal"):
        b = ctk.CTkButton(
            parent, text=text, fg_color=fg, hover_color=hover,
            font=ctk.CTkFont(size=12, weight="bold"),
            height=40, corner_radius=8, state=state, command=cmd)
        b.grid(row=row, column=0, padx=14, pady=3, sticky="ew")
        return b

    # ── Main content ──────────────────────────────────────────────────────────
    def _build_content(self):
        main = ctk.CTkFrame(self, corner_radius=0, fg_color=C_BG)
        main.grid(row=0, column=1, sticky="nsew")
        main.grid_columnconfigure(0, weight=1)
        main.grid_rowconfigure(2, weight=1)

        # ── Top bar ───────────────────────────────────────────────────────────
        top = ctk.CTkFrame(main, height=52, corner_radius=0, fg_color=C_PANEL)
        top.grid(row=0, column=0, sticky="ew")
        top.grid_columnconfigure(1, weight=1)
        top.grid_propagate(False)

        ctk.CTkLabel(top, text="Extracted Business Data",
                     font=ctk.CTkFont(size=14, weight="bold")
                     ).grid(row=0, column=0, padx=18, pady=14, sticky="w")

        self.lbl_status = ctk.CTkLabel(
            top, text="Ready — enter queries and click Start",
            font=ctk.CTkFont(size=11), text_color=C_MUTED)
        self.lbl_status.grid(row=0, column=1, padx=10, pady=14, sticky="w")

        self.lbl_count = ctk.CTkLabel(
            top, text="0 records",
            font=ctk.CTkFont(size=11), text_color=C_MUTED)
        self.lbl_count.grid(row=0, column=2, padx=18, pady=14, sticky="e")

        # ── Filter bar ────────────────────────────────────────────────────────
        fbar = ctk.CTkFrame(main, height=44, corner_radius=0, fg_color=C_CARD)
        fbar.grid(row=1, column=0, sticky="ew")
        fbar.grid_columnconfigure(1, weight=1)
        fbar.grid_propagate(False)

        ctk.CTkLabel(fbar, text="🔍", font=ctk.CTkFont(size=13)
                     ).grid(row=0, column=0, padx=(14, 4), pady=10)

        self.var_filter = tk.StringVar()
        self.var_filter.trace_add("write", lambda *_: self._apply_filter())
        ctk.CTkEntry(fbar, textvariable=self.var_filter,
                     placeholder_text="Filter results by any field…",
                     height=28, font=ctk.CTkFont(size=11),
                     border_color="#2A2A5A"
                     ).grid(row=0, column=1, padx=6, pady=8, sticky="ew")

        ctk.CTkButton(fbar, text="✕ Clear Filter", width=100, height=28,
                      fg_color="#2A2A55", hover_color="#3A3A70",
                      font=ctk.CTkFont(size=11),
                      command=lambda: self.var_filter.set("")
                      ).grid(row=0, column=2, padx=(4, 14), pady=8)

        # ── Table ─────────────────────────────────────────────────────────────
        tf = ctk.CTkFrame(main, corner_radius=0, fg_color="transparent")
        tf.grid(row=2, column=0, sticky="nsew")
        tf.grid_columnconfigure(0, weight=1)
        tf.grid_rowconfigure(0, weight=1)

        style = ttk.Style()
        style.theme_use("clam")
        style.configure("M.Treeview",
                         background="#13132B", foreground="#D8D8FF",
                         rowheight=26, fieldbackground="#13132B",
                         font=("Calibri", 10), borderwidth=0)
        style.configure("M.Treeview.Heading",
                         background="#0D47A1", foreground="white",
                         font=("Calibri", 10, "bold"), relief="flat")
        style.map("M.Treeview",
                  background=[("selected", "#1565C0")],
                  foreground=[("selected", "white")])
        style.map("M.Treeview.Heading",
                  background=[("active", C_ACCENT2)])

        self.tree = ttk.Treeview(
            tf, columns=DATA_COLS, show="headings",
            style="M.Treeview", selectmode="extended")

        widths = {
            "Name": 190, "Address": 230, "Category": 125,
            "Phone": 125, "Website": 190, "Email": 175, "Query": 155,
        }
        for col in DATA_COLS:
            self.tree.heading(col, text=col,
                              command=lambda c=col: self._sort_col(c))
            self.tree.column(col, width=widths[col], minwidth=70, stretch=True)

        vsb = ttk.Scrollbar(tf, orient="vertical",   command=self.tree.yview)
        hsb = ttk.Scrollbar(tf, orient="horizontal",  command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")

        self.tree.tag_configure("odd",  background="#13132B")
        self.tree.tag_configure("even", background="#181832")

        # ── Progress bar ──────────────────────────────────────────────────────
        pf = ctk.CTkFrame(main, height=46, corner_radius=0, fg_color=C_CARD)
        pf.grid(row=3, column=0, sticky="ew")
        pf.grid_columnconfigure(0, weight=1)
        pf.grid_propagate(False)

        self.progress = ctk.CTkProgressBar(pf, mode="indeterminate", height=5,
                                           progress_color=C_ACCENT)
        self.progress.grid(row=0, column=0, padx=14, pady=(6, 2), sticky="ew")
        self.progress.set(0)

        self.lbl_prog = ctk.CTkLabel(
            pf, text="", font=ctk.CTkFont(size=10), text_color=C_MUTED)
        self.lbl_prog.grid(row=1, column=0, padx=14, pady=(0, 6), sticky="w")

        # ── Context menu ──────────────────────────────────────────────────────
        self._ctx = tk.Menu(self, tearoff=0,
                            bg="#1A1A38", fg=C_FG,
                            activebackground=C_ACCENT, activeforeground="white")
        for label, cmd in [
            ("Copy Name",    lambda: self._copy_field("Name")),
            ("Copy Phone",   lambda: self._copy_field("Phone")),
            ("Copy Website", lambda: self._copy_field("Website")),
            ("Copy Email",   lambda: self._copy_field("Email")),
            (None, None),
            ("Copy Entire Row", self._copy_row),
            (None, None),
            ("Delete Selected", self._del_selected),
        ]:
            if label:
                self._ctx.add_command(label=label, command=cmd)
            else:
                self._ctx.add_separator()

        self.tree.bind("<Button-3>",   self._show_ctx)
        self.tree.bind("<Double-1>",   self._dbl_click)

    # ═════════════════════════════════════════════════════════════════════════
    # SCRAPER CONTROL
    # ═════════════════════════════════════════════════════════════════════════

    def _start(self):
        raw = self.txt_queries.get("1.0", "end").strip()
        queries = [q.strip() for q in raw.splitlines() if q.strip()]
        if not queries:
            messagebox.showwarning("No Queries",
                "Please enter at least one search query.")
            return

        if self._records:
            if not messagebox.askyesno("Existing Data",
                f"You have {len(self._records)} existing records.\n"
                "New results will be appended. Continue?"):
                return

        self._running = True
        self.btn_start.configure(state="disabled")
        self.btn_stop.configure(state="normal")
        self.progress.start()
        self._set_status(f"Starting for {len(queries)} "
                         f"{'query' if len(queries)==1 else 'queries'}…")

        self._scraper = MapsScraper(
            headless       = self.var_headless.get(),
            extract_emails = self.var_emails.get(),
        )
        self._thread = threading.Thread(
            target=self._scraper.run,
            args=(queries, self._on_result, self._on_complete),
            daemon=True,
        )
        self._thread.start()

    def _stop(self):
        if self._scraper:
            self._scraper.stop()
        self._set_status("Stopping…", "warning")
        self.btn_stop.configure(state="disabled")

    # ═════════════════════════════════════════════════════════════════════════
    # CALLBACKS  (thread → main)
    # ═════════════════════════════════════════════════════════════════════════

    def _on_result(self, data: dict, total: int):
        """Called from scraper thread — schedule UI update on main thread."""
        self.after(0, self._add_row, data, total)

    def _on_complete(self):
        self.after(0, self._finish)

    def _finish(self):
        self._running = False
        self.btn_start.configure(state="normal")
        self.btn_stop.configure(state="disabled")
        self.progress.stop()
        self.progress.set(0)
        if self._records:
            self.btn_export.configure(state="normal")

    # ═════════════════════════════════════════════════════════════════════════
    # QUEUE POLLING  (status messages)
    # ═════════════════════════════════════════════════════════════════════════

    def _poll(self):
        if self._scraper:
            while not self._scraper.status_q.empty():
                try:
                    item = self._scraper.status_q.get_nowait()
                    self._set_status(item["msg"], item.get("level", "info"))
                except queue.Empty:
                    break
        self.after(120, self._poll)

    # ═════════════════════════════════════════════════════════════════════════
    # TABLE HELPERS
    # ═════════════════════════════════════════════════════════════════════════

    def _add_row(self, data: dict, total: int):
        tag = "even" if total % 2 == 0 else "odd"
        vals = [data.get(c, "") for c in DATA_COLS]
        self.tree.insert("", "end", values=vals, tags=(tag,))
        self.tree.yview_moveto(1.0)

        self._records.append(data)
        self.lbl_total.configure(text=str(total))
        web_cnt = sum(1 for r in self._records if r.get("Website"))
        self.lbl_web.configure(text=str(web_cnt))
        self.lbl_count.configure(text=f"{total} record{'s' if total!=1 else ''}")
        if total > 0:
            self.btn_export.configure(state="normal")

    def _apply_filter(self):
        term = self.var_filter.get().lower()
        for iid in self.tree.get_children():
            self.tree.delete(iid)
        filtered = (
            [r for r in self._records
             if any(term in str(v).lower() for v in r.values())]
            if term else self._records
        )
        for i, r in enumerate(filtered):
            tag = "even" if i % 2 == 0 else "odd"
            self.tree.insert("", "end",
                             values=[r.get(c, "") for c in DATA_COLS],
                             tags=(tag,))
        suffix = (f" (filtered from {len(self._records)})"
                  if term and len(filtered) != len(self._records) else "")
        self.lbl_count.configure(
            text=f"{len(filtered)} record{'s' if len(filtered)!=1 else ''}{suffix}")

    def _sort_col(self, col: str):
        asc = self._sort_asc.get(col, True)
        ci  = DATA_COLS.index(col)
        items = [(self.tree.set(iid, col), iid)
                 for iid in self.tree.get_children("")]
        items.sort(key=lambda x: x[0].lower(), reverse=not asc)
        for idx, (_, iid) in enumerate(items):
            self.tree.move(iid, "", idx)
            self.tree.item(iid, tags=("even" if idx%2==0 else "odd",))
        self._sort_asc[col] = not asc

    # ═════════════════════════════════════════════════════════════════════════
    # CONTEXT MENU
    # ═════════════════════════════════════════════════════════════════════════

    def _show_ctx(self, event):
        iid = self.tree.identify_row(event.y)
        if iid:
            self.tree.selection_set(iid)
            try:
                self._ctx.tk_popup(event.x_root, event.y_root)
            finally:
                self._ctx.grab_release()

    def _copy_field(self, field: str):
        sel = self.tree.selection()
        if not sel:
            return
        ci = DATA_COLS.index(field)
        vals = self.tree.item(sel[0])["values"]
        if vals:
            self.clipboard_clear()
            self.clipboard_append(str(vals[ci]))

    def _copy_row(self):
        sel = self.tree.selection()
        if not sel:
            return
        rows = ["\t".join(str(v) for v in self.tree.item(s)["values"])
                for s in sel]
        self.clipboard_clear()
        self.clipboard_append("\n".join(rows))

    def _del_selected(self):
        sel = self.tree.selection()
        if not sel:
            return
        names = {self.tree.item(s)["values"][0] for s in sel}
        for s in sel:
            self.tree.delete(s)
        self._records = [r for r in self._records
                         if r.get("Name") not in names]
        self._refresh_stats()

    def _dbl_click(self, event):
        region = self.tree.identify("region", event.x, event.y)
        if region != "cell":
            return
        col_id = self.tree.identify_column(event.x)
        iid    = self.tree.identify_row(event.y)
        if not (col_id and iid):
            return
        ci   = int(col_id.replace("#", "")) - 1
        vals = self.tree.item(iid)["values"]
        if vals and ci < len(vals):
            self.clipboard_clear()
            self.clipboard_append(str(vals[ci]))
            self._set_status(f"Copied: {str(vals[ci])[:60]}")

    # ═════════════════════════════════════════════════════════════════════════
    # TOP-LEVEL ACTIONS
    # ═════════════════════════════════════════════════════════════════════════

    def _export(self):
        if not self._records:
            messagebox.showwarning("No Data", "Nothing to export yet.")
            return
        ts   = datetime.now().strftime("%Y%m%d_%H%M%S")
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel Workbook", "*.xlsx"), ("All files", "*.*")],
            initialfile=f"google_maps_{ts}.xlsx",
            title="Save Excel Export",
        )
        if not path:
            return
        try:
            ExcelExporter().export(self._records, path)
            if messagebox.askyesno("Export Successful",
                f"Saved {len(self._records)} records to:\n{path}\n\nOpen the file now?"):
                os.startfile(path)
        except Exception as e:
            messagebox.showerror("Export Error", str(e))

    def _clear(self):
        if self._running:
            messagebox.showwarning("Scraper Running",
                "Stop the scraper before clearing results.")
            return
        if self._records:
            if not messagebox.askyesno("Clear All",
                f"Delete all {len(self._records)} records?"):
                return
        self._records.clear()
        for iid in self.tree.get_children():
            self.tree.delete(iid)
        self._refresh_stats()
        self.btn_export.configure(state="disabled")
        self._set_status("Cleared. Ready to scrape.")

    # ═════════════════════════════════════════════════════════════════════════
    # MISC
    # ═════════════════════════════════════════════════════════════════════════

    def _set_status(self, msg: str, level: str = "info"):
        colors = {
            "info":    C_MUTED,
            "success": "#66BB6A",
            "warning": "#FFA726",
            "error":   "#EF5350",
        }
        self.lbl_status.configure(text=msg, text_color=colors.get(level, C_MUTED))
        self.lbl_prog.configure(text=msg[:110])

    def _refresh_stats(self):
        n = len(self._records)
        self.lbl_total.configure(text=str(n))
        self.lbl_web.configure(
            text=str(sum(1 for r in self._records if r.get("Website"))))
        self.lbl_count.configure(
            text=f"{n} record{'s' if n!=1 else ''}")


# ══════════════════════════════════════════════════════════════════════════════
# ENTRY POINT
# ══════════════════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    app = App()
    app.mainloop()
