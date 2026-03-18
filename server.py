#!/usr/bin/env python3
"""
Scrappy — FastAPI WebSocket Server
Serves the React frontend and handles real-time scraping over WebSocket.

Author    : Soubarna Karmakar
Copyright : © 2025 Soubarna Karmakar. All rights reserved.
Version   : 2.0
"""

__author__    = "Soubarna Karmakar"
__version__   = "2.0"

import asyncio
import io
import json
import os
import subprocess
import sys
import threading
from datetime import datetime
from pathlib import Path
from typing import Any, List, Optional

# ── Third-party (all required, fail loudly here if missing) ──────────────────
from fastapi import FastAPI, WebSocket, WebSocketDisconnect
from fastapi.responses import FileResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
import uvicorn
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

# NOTE: scraper_engine and email_extractor are imported LAZILY inside the
# WebSocket handler so that a Playwright import error does NOT crash the
# server on startup — the user still sees the UI and a friendly error message.

# ── Resolve frontend dist folder (PyInstaller-safe) ───────────────────────────
if getattr(sys, "frozen", False):
    BASE_DIR = Path(sys._MEIPASS)          # type: ignore[attr-defined]
else:
    BASE_DIR = Path(__file__).parent

FRONTEND_DIST = BASE_DIR / "frontend" / "dist"

# ── FastAPI app ────────────────────────────────────────────────────────────────
app = FastAPI(title="Scrappy", version="2.0")

# ── Server-side result store ──────────────────────────────────────────────────
# Results are stored here so the export endpoint can serve them directly
# without the browser sending a large JSON POST body (which breaks on big datasets).
_result_store: List[dict] = []
_result_lock  = threading.Lock()

# One active scraper at a time
_active_scraper: Optional[Any] = None
_scraper_lock   = threading.Lock()


# ── Excel columns ──────────────────────────────────────────────────────────────
_COLS      = ["Name", "Address", "Category", "Phone", "Website", "Email",
              "Rating", "Reviews", "Query"]
_COL_WIDTH = {"Name": 28, "Address": 38, "Category": 18,
              "Phone": 16, "Website": 32, "Email": 28,
              "Rating": 10, "Reviews": 12, "Query": 22}
_NUM_COLS   = {"Reviews"}   # stored as int in Excel
_PHONE_COLS = {"Phone"}     # @-format: preserves + and leading zeros


# ── WebSocket endpoint ─────────────────────────────────────────────────────────
@app.websocket("/ws")
async def websocket_endpoint(ws: WebSocket) -> None:
    global _active_scraper
    await ws.accept()
    loop = asyncio.get_event_loop()

    # ── helpers ────────────────────────────────────────────────────────────────
    async def send(payload: dict) -> None:
        try:
            await ws.send_text(json.dumps(payload))
        except Exception:
            pass

    def _cb_status(message: str, level: str = "info") -> None:
        asyncio.run_coroutine_threadsafe(
            send({"type": "status", "message": message, "level": level}), loop)

    def _cb_progress(current: int, total: int, query: str) -> None:
        asyncio.run_coroutine_threadsafe(
            send({"type": "progress", "current": current,
                  "total": total, "query": query}), loop)

    def _cb_complete(total: int) -> None:
        asyncio.run_coroutine_threadsafe(
            send({"type": "complete", "total": total}), loop)

    # ── message loop ──────────────────────────────────────────────────────────
    try:
        while True:
            raw  = await ws.receive_text()
            msg  = json.loads(raw)
            kind = msg.get("type")

            # START ────────────────────────────────────────────────────────────
            if kind == "start":
                queries = [q.strip() for q in msg.get("queries", []) if q.strip()]
                opts    = msg.get("options", {})

                if not queries:
                    await send({"type": "status",
                                "message": "No queries provided.", "level": "error"})
                    continue

                # Lazy import — isolated here so a bad install only shows in UI
                try:
                    from scraper_engine import ScraperEngine   # noqa: PLC0415
                    from email_extractor import EmailExtractor  # noqa: PLC0415
                except ImportError as exc:
                    await send({"type": "status",
                                "message": "Import error: %s. Re-run install.bat." % exc,
                                "level": "error"})
                    continue

                with _scraper_lock:
                    if _active_scraper and _active_scraper.is_running:
                        await send({"type": "status",
                                    "message": "Scraper already running.",
                                    "level": "warning"})
                        continue

                # Clear the store for the new run
                with _result_lock:
                    _result_store.clear()

                dedup = bool(opts.get("dedup", False))

                # Build the result callback with dedup logic
                def _make_cb_result(dedup_flag: bool):
                    def _cb_result(data: dict) -> None:
                        with _result_lock:
                            if dedup_flag:
                                name  = data.get("Name", "").strip()
                                phone = data.get("Phone", "").strip()
                                if name and phone:
                                    for i, existing in enumerate(_result_store):
                                        if (existing.get("Name", "").strip() == name and
                                                existing.get("Phone", "").strip() == phone):
                                            # Replace in store and notify frontend
                                            _result_store[i] = data
                                            asyncio.run_coroutine_threadsafe(
                                                send({"type": "result_replace",
                                                      "index": i, "data": data}), loop)
                                            return
                            _result_store.append(data)
                        asyncio.run_coroutine_threadsafe(
                            send({"type": "result", "data": data}), loop)
                    return _cb_result

                with _scraper_lock:
                    email_ext = (
                        EmailExtractor() if opts.get("extractEmails") else None
                    )
                    _active_scraper = ScraperEngine(
                        headless        = bool(opts.get("headless", False)),
                        phone_only      = bool(opts.get("phoneOnly", False)),
                        email_extractor = email_ext,
                        on_result       = _make_cb_result(dedup),
                        on_status       = _cb_status,
                        on_progress     = _cb_progress,
                        on_complete     = _cb_complete,
                    )

                threading.Thread(
                    target=_active_scraper.run,
                    args=(queries,),
                    daemon=True,
                ).start()

            # STOP ─────────────────────────────────────────────────────────────
            elif kind == "stop":
                with _scraper_lock:
                    if _active_scraper:
                        _active_scraper.stop()
                await send({"type": "status",
                            "message": "Stop signal sent.", "level": "warning"})

            # CLEAR ────────────────────────────────────────────────────────────
            elif kind == "clear":
                with _result_lock:
                    _result_store.clear()

            # PING ─────────────────────────────────────────────────────────────
            elif kind == "ping":
                await send({"type": "pong"})

    except WebSocketDisconnect:
        with _scraper_lock:
            if _active_scraper:
                _active_scraper.stop()
    except Exception as exc:
        try:
            await send({"type": "status",
                        "message": "Server error: %s" % exc, "level": "error"})
        except Exception:
            pass


# ── Shared Excel builder ──────────────────────────────────────────────────────
def _build_workbook(rows_snapshot: list) -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Scrappy Data"

    hdr_font  = Font(name="Calibri", bold=True, size=11)
    hdr_align = Alignment(horizontal="center", vertical="center")

    for ci, col in enumerate(_COLS, 1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.font      = hdr_font
        cell.alignment = hdr_align
        ws.column_dimensions[get_column_letter(ci)].width = _COL_WIDTH.get(col, 20)

    for ri, row in enumerate(rows_snapshot, 2):
        for ci, col in enumerate(_COLS, 1):
            raw = row.get(col, "") or ""
            if col in _NUM_COLS and raw:
                try:
                    raw = int(raw)
                except (ValueError, TypeError):
                    pass
            elif col == "Rating" and raw:
                try:
                    raw = float(raw)
                except (ValueError, TypeError):
                    pass
            cell = ws.cell(row=ri, column=ci, value=raw)
            cell.alignment = Alignment(vertical="center")
            if col in _PHONE_COLS and raw:
                cell.number_format = "@"   # text format — keeps + and leading zeros

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = (
        "A1:%s%d" % (get_column_letter(len(_COLS)), len(rows_snapshot) + 1)
    )
    ws.row_dimensions[1].height = 22

    ws2 = wb.create_sheet("Summary")
    ws2["A1"] = "Scrappy — Export Summary"
    ws2["A1"].font = Font(bold=True, size=13)
    ws2["A3"] = "Author";    ws2["B3"] = "Soubarna Karmakar"
    ws2["A4"] = "Generated"; ws2["B4"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws2["A5"] = "Records";   ws2["B5"] = len(rows_snapshot)
    ws2.column_dimensions["A"].width = 16
    ws2.column_dimensions["B"].width = 30

    return wb


# ── Excel export: save to disk + open (desktop app — no browser download) ─────
@app.get("/export-open")
async def export_open() -> dict:
    """
    Save the Excel file to the user's Downloads folder (or next to the exe
    if Downloads is unavailable) and open it with the OS default application.
    Returns JSON so the frontend can show a success/error message.
    """
    with _result_lock:
        rows_snapshot = list(_result_store)

    if not rows_snapshot:
        return {"status": "error", "message": "No data to export."}

    ts       = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = "scrappy_%s.xlsx" % ts

    # Prefer ~/Downloads; fall back to the executable's directory
    downloads = Path.home() / "Downloads"
    if not downloads.exists():
        downloads = Path(sys.executable).parent if getattr(sys, "frozen", False) \
                    else Path(__file__).parent
    out_path = downloads / filename

    try:
        wb = _build_workbook(rows_snapshot)
        wb.save(str(out_path))
    except Exception as exc:
        return {"status": "error", "message": "Failed to save file: %s" % exc}

    # Open the file with the OS default app (Excel on Windows)
    try:
        if sys.platform == "win32":
            os.startfile(str(out_path))           # type: ignore[attr-defined]
        elif sys.platform == "darwin":
            subprocess.Popen(["open", str(out_path)])
        else:
            subprocess.Popen(["xdg-open", str(out_path)])
    except Exception:
        pass  # opening failed but file was saved — still report success

    return {"status": "ok", "path": str(out_path), "records": len(rows_snapshot)}


# ── Results snapshot (lets the frontend rehydrate after a page refresh) ───────
@app.get("/results")
async def get_results() -> dict:
    with _result_lock:
        return {"results": list(_result_store)}


# ── Excel export (streaming — kept for browser / dev usage) ───────────────────
@app.get("/export")
async def export_excel() -> StreamingResponse:
    with _result_lock:
        rows_snapshot = list(_result_store)

    wb  = _build_workbook(rows_snapshot)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    # NOTE: do NOT call wb.close() — not available in all openpyxl versions
    # and calling it after save() serves no purpose (buf is already populated).

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="scrappy_%s.xlsx"' % ts},
    )


# ── Serve React build (SPA fallback) ──────────────────────────────────────────
if FRONTEND_DIST.exists():
    _assets_dir = FRONTEND_DIST / "assets"
    if _assets_dir.exists():
        app.mount("/assets",
                  StaticFiles(directory=str(_assets_dir)),
                  name="assets")

    @app.get("/{full_path:path}")
    async def spa_fallback(full_path: str) -> FileResponse:
        return FileResponse(str(FRONTEND_DIST / "index.html"))

else:
    @app.get("/")
    async def no_frontend() -> dict:
        return {
            "status": "error",
            "message": "Frontend not built.",
            "fix": "cd frontend && npm install && npm run build",
        }


# ── Standalone entry (dev mode: python server.py) ─────────────────────────────
def start_server(host: str = "127.0.0.1", port: int = 7410) -> None:
    # log_config=None disables uvicorn's default logging setup which references
    # a 'default' formatter that is unavailable in PyInstaller frozen bundles.
    uvicorn.run(app, host=host, port=port, log_level="warning", log_config=None)


if __name__ == "__main__":
    start_server()
